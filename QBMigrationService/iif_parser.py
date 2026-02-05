"""
QuickBooks Desktop IIF File Parser
Parses IIF (Intuit Interchange Format) files exported from QuickBooks Desktop

IIF is a tab-delimited format used by QuickBooks for importing/exporting data.
This parser converts IIF files to JSON format for migration to QuickBooks Online.
"""

import csv
import json
import logging
import os
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from io import StringIO
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)


def safe_decimal(value, default: str = "0") -> Decimal:
    """
    CRITICAL FIX: Safely convert value to Decimal, handling None and invalid inputs.

    Args:
        value: Value to convert (can be None, str, int, float, Decimal)
        default: Default value if conversion fails

    Returns:
        Decimal representation preserving precision
    """
    if value is None or value == "":
        return Decimal(default)
    if isinstance(value, Decimal):
        return value
    try:
        # Convert to string first to preserve precision
        return Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError) as e:
        logger.warning(
            f"Failed to convert '{value}' to Decimal: {e}, using default '{default}'"
        )
        return Decimal(default)


class IIFParser:
    """Parser for QuickBooks Desktop IIF export files"""

    # IIF record types and their field mappings
    RECORD_TYPES = {
        "CUST": "customers",
        "VEND": "vendors",
        "INVITEM": "items",
        "ACCNT": "accounts",
        "EMP": "employees",
        "TRNS": "transactions",
        "SPL": "transaction_splits",
        "CLASS": "classes",
        "TERMS": "terms",
        "SHIPMETH": "shipping_methods",
        "PAYMETH": "payment_methods",
        "OTHERNAME": "other_names",
        "INVMEMO": "memos",
        "TODO": "todos",
        "TIMEACT": "time_activities",
        "BUDGET": "budgets",
    }

    def __init__(self):
        self.data = {key: [] for key in self.RECORD_TYPES.values()}
        self.headers = {}
        self.errors = []
        self._current_transaction = None
        self.stats = {
            "total_lines": 0,
            "parsed_records": 0,
            "skipped_lines": 0,
            "errors": 0,
        }

    def parse_file(self, file_path: str) -> Dict[str, Any]:
        """Parse an IIF file and return structured data
        AUDIT FIX: Added robust path validation with proper containment check
        """
        # Reset state for fresh parse
        self.data = {key: [] for key in self.RECORD_TYPES.values()}
        self.headers = {}
        self.errors = []
        self.stats = {
            "total_lines": 0,
            "parsed_records": 0,
            "skipped_lines": 0,
            "errors": 0,
        }

        # Security: Robust path traversal prevention
        # 1. Normalize the path to resolve any relative components
        real_path = os.path.realpath(file_path)

        # 2. Define allowed base directories (current working directory or explicit data dir)
        allowed_base_dirs = [
            os.path.realpath(os.getcwd()),
            os.path.realpath(os.path.dirname(__file__)),
        ]

        # Add DATA_DIR from environment if configured
        data_dir = os.environ.get("DATA_DIR")
        if data_dir:
            allowed_base_dirs.append(os.path.realpath(data_dir))

        # 3. Check that the resolved path is within one of the allowed directories
        is_path_allowed = any(
            real_path.startswith(base_dir + os.sep) or real_path == base_dir
            for base_dir in allowed_base_dirs
        )

        if not is_path_allowed:
            raise ValueError(
                f"Security: File path is outside allowed directories: {file_path}"
            )

        # 4. Verify the file exists and is a regular file (not symlink to outside)
        if not os.path.isfile(real_path):
            raise ValueError(f"Invalid file path (not a file): {file_path}")

        # FIX #10: Check file size before reading (prevent memory exhaustion)
        MAX_IIF_SIZE = 100 * 1024 * 1024  # 100MB limit
        file_size = os.path.getsize(real_path)
        if file_size > MAX_IIF_SIZE:
            raise ValueError(
                f"IIF file too large: {file_size / (1024*1024):.1f}MB exceeds "
                f"{MAX_IIF_SIZE / (1024*1024):.0f}MB limit. "
                f"Please split the file into smaller chunks."
            )
        logger.info(f"Reading IIF file: {file_size / 1024:.1f}KB")

        try:
            with open(real_path, "r", encoding="cp1252") as f:
                content = f.read()
        except UnicodeDecodeError:
            with open(real_path, "r", encoding="utf-8", errors="replace") as f:
                content = f.read()
        return self.parse_content(content)

    def parse_content(self, content: str) -> Dict[str, Any]:
        """Parse IIF content string"""
        # Reset state for fresh parse
        self.data = {key: [] for key in self.RECORD_TYPES.values()}
        self.headers = {}
        self.errors = []
        self._current_transaction = None
        self.stats = {
            "total_lines": 0,
            "parsed_records": 0,
            "skipped_lines": 0,
            "errors": 0,
        }

        lines = content.replace("\r\n", "\n").replace("\r", "\n").split("\n")
        self.stats["total_lines"] = len(lines)

        # FIX #11: Limit line count to prevent pathological inputs from hanging
        MAX_LINES = 1_000_000  # 1 million line limit
        if len(lines) > MAX_LINES:
            raise ValueError(
                f"IIF file has too many lines: {len(lines):,} exceeds {MAX_LINES:,} limit. "
                f"Please split the file into smaller chunks."
            )
        logger.info(f"Parsing IIF content: {len(lines):,} lines")

        # AUDIT FIX: Removed unused current_type variable

        for line_num, line in enumerate(lines, 1):
            if not line.strip():
                self.stats["skipped_lines"] += 1
                continue

            try:
                fields = line.split("\t")
                record_type = fields[0].upper() if fields else ""

                # Header line defines columns
                if record_type.startswith("!"):
                    header_type = record_type[1:]  # Remove the !
                    self.headers[header_type] = [h.strip().upper() for h in fields[1:]]

                # Transaction line - start a new transaction block
                elif record_type == "TRNS":
                    headers = self.headers.get("TRNS", [])
                    record = {}
                    for i, value in enumerate(fields[1:], 0):
                        if i < len(headers):
                            record[headers[i]] = value.strip()
                        else:
                            record[f"field_{i}"] = value.strip()
                    record["_source_line"] = line_num
                    record["_record_type"] = "TRNS"
                    self._current_transaction = record
                    self._current_transaction["_splits"] = []
                    self.stats["parsed_records"] += 1

                # Transaction split line
                elif record_type == "SPL":
                    headers = self.headers.get("SPL", [])
                    record = {}
                    for i, value in enumerate(fields[1:], 0):
                        if i < len(headers):
                            record[headers[i]] = value.strip()
                        else:
                            record[f"field_{i}"] = value.strip()
                    record["_source_line"] = line_num
                    record["_record_type"] = "SPL"
                    if self._current_transaction is not None:
                        self._current_transaction["_splits"].append(record)
                    self.data["transaction_splits"].append(record)
                    self.stats["parsed_records"] += 1

                # End of transaction block - commit the grouped transaction
                elif record_type == "ENDTRNS":
                    if self._current_transaction is not None:
                        self.data["transactions"].append(self._current_transaction)
                        self._current_transaction = None

                # Data line (all other record types)
                elif record_type in self.RECORD_TYPES:
                    data_key = self.RECORD_TYPES[record_type]
                    headers = self.headers.get(record_type, [])

                    record = {}
                    for i, value in enumerate(fields[1:], 0):
                        if i < len(headers):
                            record[headers[i]] = value.strip()
                        else:
                            record[f"field_{i}"] = value.strip()

                    record["_source_line"] = line_num
                    record["_record_type"] = record_type
                    self.data[data_key].append(record)
                    self.stats["parsed_records"] += 1

                else:
                    self.stats["skipped_lines"] += 1

            except Exception as e:
                self.errors.append(
                    {"line": line_num, "content": line[:100], "error": str(e)}
                )
                self.stats["errors"] += 1

        return self.get_result()

    def get_result(self) -> Dict[str, Any]:
        """Get parsed result with metadata"""
        return {
            "metadata": {
                "parsed_at": datetime.now(timezone.utc).isoformat(),
                "stats": self.stats,
                "errors": self.errors[:100],  # Limit error list
            },
            "data": self.data,
            "summary": self._get_summary(),
        }

    def _get_summary(self) -> Dict[str, int]:
        """Get count of records by type"""
        return {key: len(value) for key, value in self.data.items() if value}


class QuickBooksExportParser:
    """
    Unified parser for QuickBooks Desktop exports
    Supports: IIF, CSV, Excel exports
    """

    def __init__(self):
        self.iif_parser = IIFParser()
        self.data = {}
        self.file_type = None

    def parse(self, file_path: str) -> Dict[str, Any]:
        """Parse a QuickBooks export file (auto-detect format)"""
        ext = os.path.splitext(file_path)[1].lower()

        if ext == ".iif":
            self.file_type = "iif"
            return self.iif_parser.parse_file(file_path)
        elif ext == ".csv":
            self.file_type = "csv"
            return self._parse_csv(file_path)
        elif ext in [".xls", ".xlsx"]:
            self.file_type = "excel"
            return self._parse_excel(file_path)
        else:
            raise ValueError(f"Unsupported file format: {ext}")

    def _validate_file_path(self, file_path: str) -> str:
        """Validate file path for security (path traversal prevention)"""
        real_path = os.path.realpath(file_path)
        allowed_base_dirs = [
            os.path.realpath(os.getcwd()),
            os.path.realpath(os.path.dirname(__file__)),
        ]
        data_dir = os.environ.get("DATA_DIR")
        if data_dir:
            allowed_base_dirs.append(os.path.realpath(data_dir))

        is_path_allowed = any(
            real_path.startswith(base_dir + os.sep) or real_path == base_dir
            for base_dir in allowed_base_dirs
        )
        if not is_path_allowed:
            raise ValueError(
                f"Security: File path is outside allowed directories: {file_path}"
            )
        if not os.path.isfile(real_path):
            raise ValueError(f"Invalid file path (not a file): {file_path}")
        return real_path

    def _parse_csv(self, file_path: str) -> Dict[str, Any]:
        """Parse CSV export from QuickBooks"""
        # Security: Validate path
        real_path = self._validate_file_path(file_path)
        records = []

        with open(real_path, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                # Clean up keys and values (handle None keys from extra columns)
                clean_row = {
                    (k.strip() if k else "_unknown"): v.strip() if v else ""
                    for k, v in row.items()
                    if k is not None
                }
                records.append(clean_row)

        # Detect entity type from headers or filename
        entity_type = self._detect_entity_type(file_path, records)

        return {
            "metadata": {
                "parsed_at": datetime.now(timezone.utc).isoformat(),
                "file_type": "csv",
                "record_count": len(records),
            },
            "data": {entity_type: records},
            "summary": {entity_type: len(records)},
        }

    def _parse_excel(self, file_path: str) -> Dict[str, Any]:
        """Parse Excel export from QuickBooks"""
        # Security: Validate path
        real_path = self._validate_file_path(file_path)
        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "openpyxl required for Excel parsing. Install with: pip install openpyxl"
            )

        wb = openpyxl.load_workbook(real_path, read_only=True)
        all_data = {}

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))

            if not rows:
                continue

            headers = [
                str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])
            ]
            records = []

            for row in rows[1:]:
                record = {}
                for i, value in enumerate(row):
                    if i < len(headers):
                        record[headers[i]] = str(value).strip() if value else ""
                records.append(record)

            if records:
                all_data[sheet_name.lower().replace(" ", "_")] = records

        return {
            "metadata": {
                "parsed_at": datetime.now(timezone.utc).isoformat(),
                "file_type": "excel",
                "sheets": list(all_data.keys()),
            },
            "data": all_data,
            "summary": {k: len(v) for k, v in all_data.items()},
        }

    def _detect_entity_type(self, file_path: str, records: List[Dict]) -> str:
        """Detect entity type from filename or content"""
        filename = os.path.basename(file_path).lower()

        # Check filename with word boundary matching
        type_hints = {
            "customers": "customers",
            "customer": "customers",
            "vendors": "vendors",
            "vendor": "vendors",
            "invoices": "invoices",
            "invoice": "invoices",
            "bills": "bills",
            "bill": "bills",
            "payments": "payments",
            "payment": "payments",
            "accounts": "accounts",
            "account": "accounts",
            "items": "items",
            "item": "items",
            "employees": "employees",
            "employee": "employees",
            "journal": "journal_entries",
        }

        # Use word boundary matching instead of substring
        filename_lower = (
            os.path.basename(filename)
            .replace(".csv", "")
            .replace(".xlsx", "")
            .replace(".xls", "")
        )
        for hint, entity_type in type_hints.items():
            # Check for exact match or word boundary
            if (
                filename_lower == hint
                or filename_lower.startswith(hint + "_")
                or filename_lower.endswith("_" + hint)
            ):
                return entity_type

        # Check headers
        if records:
            headers = set(records[0].keys())
            if "Customer" in headers or "Company" in headers:
                return "customers"
            elif "Vendor" in headers:
                return "vendors"
            elif "Invoice #" in headers or "Invoice Number" in headers:
                return "invoices"

        return "unknown"


def _parse_address_line3(baddr3: str) -> dict:
    """Parse IIF BADDR3 'City, State ZIP' into components."""
    import re

    result = {"City": "", "CountrySubDivisionCode": "", "PostalCode": ""}
    if not baddr3:
        return result
    # Try "City, State ZIP" format
    match = re.match(r"^(.+?),\s*(\w{2})\s+(\d{5}(?:-\d{4})?)$", baddr3)
    if match:
        result["City"] = match.group(1).strip()
        result["CountrySubDivisionCode"] = match.group(2)
        result["PostalCode"] = match.group(3)
    else:
        # Try "City, State" without ZIP
        match2 = re.match(r"^(.+?),\s*(.+)$", baddr3)
        if match2:
            result["City"] = match2.group(1).strip()
            result["CountrySubDivisionCode"] = match2.group(2).strip()
        else:
            result["City"] = baddr3  # Fall back to entire string as city
    return result


def transform_for_qbo(iif_data: Dict[str, Any]) -> Dict[str, Any]:
    """
    Transform parsed IIF data to QuickBooks Online format
    """
    qbo_data = {
        "customers": [],
        "vendors": [],
        "items": [],
        "accounts": [],
        "invoices": [],
        "bills": [],
        "payments": [],
    }

    # Transform customers
    for cust in iif_data.get("data", {}).get("customers", []):
        addr3_parts = _parse_address_line3(cust.get("BADDR3", ""))
        qbo_data["customers"].append(
            {
                "DisplayName": cust.get("NAME", ""),
                "CompanyName": cust.get("COMPANYNAME", ""),
                "GivenName": cust.get("FIRSTNAME", ""),
                "FamilyName": cust.get("LASTNAME", ""),
                "PrimaryEmailAddr": {"Address": cust.get("EMAIL", "")},
                "PrimaryPhone": {"FreeFormNumber": cust.get("PHONE1", "")},
                "BillAddr": {
                    "Line1": cust.get("BADDR1", ""),
                    "Line2": cust.get("BADDR2", ""),
                    "City": addr3_parts["City"],
                    "CountrySubDivisionCode": addr3_parts["CountrySubDivisionCode"],
                    "PostalCode": addr3_parts["PostalCode"],
                    "Country": cust.get("BADDR4", ""),
                },
                "_source": "iif_import",
                "_original_id": cust.get("NAME", ""),
            }
        )

    # Transform vendors
    for vend in iif_data.get("data", {}).get("vendors", []):
        addr3_parts = _parse_address_line3(vend.get("BADDR3", ""))
        qbo_data["vendors"].append(
            {
                "DisplayName": vend.get("NAME", ""),
                "CompanyName": vend.get("COMPANYNAME", ""),
                "GivenName": vend.get("FIRSTNAME", ""),
                "FamilyName": vend.get("LASTNAME", ""),
                "PrimaryEmailAddr": {"Address": vend.get("EMAIL", "")},
                "PrimaryPhone": {"FreeFormNumber": vend.get("PHONE1", "")},
                "BillAddr": {
                    "Line1": vend.get("BADDR1", ""),
                    "Line2": vend.get("BADDR2", ""),
                    "City": addr3_parts["City"],
                    "CountrySubDivisionCode": addr3_parts["CountrySubDivisionCode"],
                    "PostalCode": addr3_parts["PostalCode"],
                    "Country": vend.get("BADDR4", ""),
                },
                "_source": "iif_import",
                "_original_id": vend.get("NAME", ""),
            }
        )

    # Transform accounts
    # FIX #16: Complete account type mapping with logging for unmapped types
    acct_type_map = {
        "BANK": "Bank",
        "AR": "AccountsReceivable",
        "OCASSET": "OtherCurrentAsset",
        "FIXASSET": "FixedAsset",
        "OASSET": "OtherAsset",
        "AP": "AccountsPayable",
        "CCARD": "CreditCard",
        "OCURLIAB": "OtherCurrentLiability",
        "LTLIAB": "LongTermLiability",
        "EQUITY": "Equity",
        "INC": "Income",
        "OINC": "OtherIncome",
        "OTHERINC": "OtherIncome",
        "COGS": "CostOfGoodsSold",
        "EXP": "Expense",
        "OEXP": "OtherExpense",
        "OTHEREXP": "OtherExpense",
        "EXINC": "OtherIncome",
        "EXEXP": "OtherExpense",
        "NONPOSTING": "NonPosting",
        # Additional QB Desktop types
        "ASSET": "OtherAsset",
        "LIABILITY": "OtherCurrentLiability",
        "INCOME": "Income",
        "EXPENSE": "Expense",
        "CURASS": "OtherCurrentAsset",
        "CURLIAB": "OtherCurrentLiability",
    }

    unmapped_account_types = set()
    for acct in iif_data.get("data", {}).get("accounts", []):
        original_type = acct.get("ACCNTTYPE", "")
        qbo_type = acct_type_map.get(original_type)

        if qbo_type is None:
            # FIX #16: Log unmapped types and use smarter default
            unmapped_account_types.add(original_type)
            # Default based on type name heuristics
            if "ASSET" in original_type.upper():
                qbo_type = "OtherAsset"
            elif "LIAB" in original_type.upper():
                qbo_type = "OtherCurrentLiability"
            elif "INC" in original_type.upper() or "REV" in original_type.upper():
                qbo_type = "Income"
            elif "EQUITY" in original_type.upper():
                qbo_type = "Equity"
            else:
                qbo_type = "Expense"  # Last resort default

        qbo_data["accounts"].append(
            {
                "Name": acct.get("NAME", ""),
                "AccountType": qbo_type,
                "Description": acct.get("DESC", ""),
                "_source": "iif_import",
                "_original_id": acct.get("NAME", ""),
                "_original_type": original_type,  # Preserve original for debugging
            }
        )

    # Log warning for unmapped account types
    if unmapped_account_types:
        logger.warning(
            f"FIX #16: {len(unmapped_account_types)} unmapped account types found: "
            f"{sorted(unmapped_account_types)}. Please verify mappings are correct."
        )

    # Transform items
    for item in iif_data.get("data", {}).get("items", []):
        item_type = item.get("INVITEMTYPE", "SERVICE")
        type_map = {
            "SERVICE": "Service",
            "SERV": "Service",
            "INVENTORY": "Inventory",
            "NON-INVENTORY": "NonInventory",
            "PART": "NonInventory",
            "OTHCHARGE": "Service",
            "SUBTOTAL": "Service",
            "DISCOUNT": "Service",
            "GROUP": "Group",
            "ASSEMBLY": "Inventory",
            "PAYMENT": "Service",
            "TAX": "Service",
            "FIXEDASSET": "NonInventory",
        }
        qbo_type = type_map.get(item_type, "Service")
        qbo_data["items"].append(
            {
                "Name": item.get("NAME", ""),
                "Description": item.get("DESC", ""),
                "UnitPrice": safe_decimal(item.get("PRICE")),
                "Type": qbo_type,
                "_source": "iif_import",
                "_original_id": item.get("NAME", ""),
            }
        )

    return qbo_data


# CLI for testing
if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        logger.info("Usage: python iif_parser.py <file.iif>")
        sys.exit(1)

    parser = QuickBooksExportParser()
    result = parser.parse(sys.argv[1])

    logger.info(json.dumps(result, indent=2, default=str))
